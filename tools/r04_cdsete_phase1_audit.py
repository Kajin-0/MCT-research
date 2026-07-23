#!/usr/bin/env python3
"""Acquire and structurally audit the public CdSeTe PL source-data archive.

This tool is intentionally fail-closed.  It records archive provenance and file
structure without interpreting a rendered figure as source-native data and
without unpickling untrusted objects.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import mimetypes
import os
from pathlib import Path
import shutil
import sys
import time
import urllib.request
import zipfile
from typing import Any

DATASET_URL = "https://zenodo.org/records/13869384/files/Datasets.zip?download=1"
DATASET_DOI = "10.5281/zenodo.13869384"
ARTICLE_DOI = "10.1038/s41467-024-52889-z"
EXPECTED_MD5 = "1401ee9b5372edb78f888d152940fc79"

TEXT_SUFFIXES = {".txt", ".csv", ".tsv", ".dat", ".asc", ".md"}
IMAGE_SUFFIXES = {".tif", ".tiff", ".png", ".jpg", ".jpeg", ".bmp"}
ARRAY_SUFFIXES = {".npy", ".npz", ".mat", ".h5", ".hdf5", ".hdf"}
SPATIAL_KEYWORDS = {
    "x",
    "y",
    "position",
    "coordinate",
    "pixel",
    "map",
    "image",
    "scan",
    "wavelength",
    "energy",
    "peak",
    "fwhm",
    "linewidth",
    "plqy",
    "intensity",
    "spectrum",
    "spectra",
}


def digest(path: Path, algorithm: str) -> str:
    h = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def download(url: str, destination: Path, attempts: int = 4) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "MCT-research-R04-public-data-audit/1.0",
            "Accept": "application/zip,application/octet-stream;q=0.9,*/*;q=0.1",
        },
    )
    last_error: Exception | None = None
    for attempt in range(1, attempts + 1):
        partial = destination.with_suffix(destination.suffix + ".partial")
        partial.unlink(missing_ok=True)
        try:
            with urllib.request.urlopen(request, timeout=90) as response, partial.open("wb") as out:
                shutil.copyfileobj(response, out, length=1024 * 1024)
            if partial.stat().st_size == 0:
                raise RuntimeError("download returned an empty file")
            partial.replace(destination)
            return
        except Exception as exc:  # pragma: no cover - network-dependent
            last_error = exc
            partial.unlink(missing_ok=True)
            if attempt < attempts:
                time.sleep(2**attempt)
    raise RuntimeError(f"failed to download {url!r} after {attempts} attempts") from last_error


def safe_extract(archive: Path, destination: Path) -> list[dict[str, Any]]:
    destination.mkdir(parents=True, exist_ok=True)
    root = destination.resolve()
    records: list[dict[str, Any]] = []
    with zipfile.ZipFile(archive) as zf:
        for info in zf.infolist():
            target = (destination / info.filename).resolve()
            if target != root and root not in target.parents:
                raise RuntimeError(f"unsafe archive path: {info.filename}")
            records.append(
                {
                    "name": info.filename,
                    "is_directory": info.is_dir(),
                    "compressed_size": info.compress_size,
                    "uncompressed_size": info.file_size,
                    "crc32": f"{info.CRC:08x}",
                }
            )
        zf.extractall(destination)
    return records


def sanitize(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, str)):
        return value
    if isinstance(value, float):
        if value != value:
            return "NaN"
        if value == float("inf"):
            return "Infinity"
        if value == float("-inf"):
            return "-Infinity"
        return value
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(k): sanitize(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [sanitize(v) for v in value]
    if hasattr(value, "item"):
        try:
            return sanitize(value.item())
        except Exception:
            pass
    return repr(value)


def inspect_text(path: Path) -> dict[str, Any]:
    raw = path.read_bytes()
    text = raw.decode("utf-8", errors="replace")
    lines = text.splitlines()
    result: dict[str, Any] = {
        "line_count": len(lines),
        "replacement_character_count": text.count("\ufffd"),
        "preview": lines[:8],
    }
    if not lines:
        return result

    sample = "\n".join(lines[:50])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        reader = csv.reader(lines, dialect)
        rows = list(reader)
        width = max((len(row) for row in rows), default=0)
        result["delimited_table"] = {
            "delimiter": dialect.delimiter,
            "row_count": len(rows),
            "maximum_column_count": width,
            "header": rows[0] if rows else [],
            "rectangular": bool(rows) and all(len(row) == width for row in rows),
        }
    except csv.Error:
        result["delimited_table"] = None
    return result


def inspect_numpy(path: Path) -> dict[str, Any]:
    import numpy as np

    loaded = np.load(path, allow_pickle=False)
    if isinstance(loaded, np.lib.npyio.NpzFile):
        arrays = {}
        for key in loaded.files:
            array = loaded[key]
            arrays[key] = {
                "shape": list(array.shape),
                "dtype": str(array.dtype),
                "finite_fraction": float(np.isfinite(array).mean()) if np.issubdtype(array.dtype, np.number) else None,
            }
        loaded.close()
        return {"container": "npz", "arrays": arrays}
    array = loaded
    return {
        "container": "npy",
        "shape": list(array.shape),
        "dtype": str(array.dtype),
        "finite_fraction": float(np.isfinite(array).mean()) if np.issubdtype(array.dtype, np.number) else None,
    }


def inspect_mat(path: Path) -> dict[str, Any]:
    try:
        from scipy.io import whosmat

        variables = [
            {"name": name, "shape": list(shape), "class": class_name}
            for name, shape, class_name in whosmat(path)
        ]
        return {"format": "matlab", "variables": variables}
    except Exception as scipy_error:
        try:
            import h5py

            return {
                "format": "matlab_hdf5",
                "datasets": inspect_hdf5(path)["datasets"],
                "scipy_error": repr(scipy_error),
            }
        except Exception as hdf_error:
            return {
                "format": "unreadable_mat",
                "scipy_error": repr(scipy_error),
                "hdf5_error": repr(hdf_error),
            }


def inspect_hdf5(path: Path) -> dict[str, Any]:
    import h5py

    datasets: list[dict[str, Any]] = []
    attrs: dict[str, Any] = {}
    with h5py.File(path, "r") as handle:
        attrs = {str(k): sanitize(v) for k, v in handle.attrs.items()}

        def visitor(name: str, obj: Any) -> None:
            if isinstance(obj, h5py.Dataset):
                datasets.append(
                    {
                        "name": name,
                        "shape": list(obj.shape),
                        "dtype": str(obj.dtype),
                        "attributes": {str(k): sanitize(v) for k, v in obj.attrs.items()},
                    }
                )

        handle.visititems(visitor)
    return {"format": "hdf5", "attributes": attrs, "datasets": datasets}


def inspect_tiff(path: Path) -> dict[str, Any]:
    import tifffile

    with tifffile.TiffFile(path) as tif:
        return {
            "series": [
                {
                    "shape": list(series.shape),
                    "dtype": str(series.dtype),
                    "axes": series.axes,
                }
                for series in tif.series
            ],
            "page_count": len(tif.pages),
            "imagej_metadata": sanitize(tif.imagej_metadata),
            "ome_metadata_present": tif.ome_metadata is not None,
        }


def inspect_image(path: Path) -> dict[str, Any]:
    from PIL import Image

    with Image.open(path) as image:
        return {
            "format": image.format,
            "size": list(image.size),
            "mode": image.mode,
            "frame_count": getattr(image, "n_frames", 1),
        }


def inspect_excel(path: Path) -> dict[str, Any]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        preview: list[list[Any]] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index <= 8:
                preview.append([sanitize(cell) for cell in row])
            if row_index > 8:
                break
        sheets.append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "preview": preview,
            }
        )
    workbook.close()
    return {"sheets": sheets}


def inspect_nested_zip(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        return {
            "member_count": len(archive.infolist()),
            "members": [
                {
                    "name": item.filename,
                    "compressed_size": item.compress_size,
                    "uncompressed_size": item.file_size,
                }
                for item in archive.infolist()
            ],
        }


def inspect_file(path: Path, root: Path) -> dict[str, Any]:
    relative = path.relative_to(root).as_posix()
    suffix = path.suffix.lower()
    record: dict[str, Any] = {
        "path": relative,
        "size_bytes": path.stat().st_size,
        "sha256": digest(path, "sha256"),
        "suffix": suffix,
        "mime_guess": mimetypes.guess_type(path.name)[0],
    }
    try:
        if suffix in TEXT_SUFFIXES:
            record["inspection"] = inspect_text(path)
        elif suffix in {".npy", ".npz"}:
            record["inspection"] = inspect_numpy(path)
        elif suffix == ".mat":
            record["inspection"] = inspect_mat(path)
        elif suffix in {".h5", ".hdf5", ".hdf"}:
            record["inspection"] = inspect_hdf5(path)
        elif suffix in {".tif", ".tiff"}:
            record["inspection"] = inspect_tiff(path)
        elif suffix in IMAGE_SUFFIXES:
            record["inspection"] = inspect_image(path)
        elif suffix in {".xlsx", ".xlsm"}:
            record["inspection"] = inspect_excel(path)
        elif suffix == ".zip":
            record["inspection"] = inspect_nested_zip(path)
        elif suffix in {".pkl", ".pickle", ".joblib"}:
            record["inspection"] = {
                "status": "not_opened",
                "reason": "untrusted serialized object; structural audit does not unpickle",
            }
        else:
            with path.open("rb") as handle:
                signature = handle.read(32)
            record["inspection"] = {"first_32_bytes_hex": signature.hex()}
    except Exception as exc:
        record["inspection_error"] = repr(exc)
    return record


def find_shapes(value: Any, location: str = "") -> list[dict[str, Any]]:
    found: list[dict[str, Any]] = []
    if isinstance(value, dict):
        for key, child in value.items():
            child_location = f"{location}.{key}" if location else str(key)
            if key == "shape" and isinstance(child, list) and all(isinstance(x, int) for x in child):
                found.append({"location": child_location, "shape": child})
            found.extend(find_shapes(child, child_location))
    elif isinstance(value, list):
        for index, child in enumerate(value):
            found.extend(find_shapes(child, f"{location}[{index}]"))
    return found


def keyword_hits(record: dict[str, Any]) -> list[str]:
    blob = json.dumps(record, sort_keys=True).lower()
    return sorted(keyword for keyword in SPATIAL_KEYWORDS if keyword in blob)


def build_summary(
    archive_path: Path,
    archive_members: list[dict[str, Any]],
    file_records: list[dict[str, Any]],
) -> dict[str, Any]:
    extension_counts: dict[str, int] = {}
    candidate_arrays: list[dict[str, Any]] = []
    for record in file_records:
        suffix = record["suffix"] or "<none>"
        extension_counts[suffix] = extension_counts.get(suffix, 0) + 1
        hits = keyword_hits(record)
        for shape_record in find_shapes(record.get("inspection", {})):
            shape = shape_record["shape"]
            if len(shape) >= 2 and sum(d > 3 for d in shape) >= 2:
                candidate_arrays.append(
                    {
                        "path": record["path"],
                        "shape_location": shape_record["location"],
                        "shape": shape,
                        "keyword_hits": hits,
                    }
                )
        table = record.get("inspection", {}).get("delimited_table")
        if isinstance(table, dict) and table.get("row_count", 0) > 3 and table.get("maximum_column_count", 0) > 3:
            candidate_arrays.append(
                {
                    "path": record["path"],
                    "shape_location": "inspection.delimited_table",
                    "shape": [table["row_count"], table["maximum_column_count"]],
                    "keyword_hits": hits,
                    "caution": "two-dimensional table; spatial meaning not established structurally",
                }
            )

    return {
        "schema_version": "1.0",
        "dataset": {
            "record_doi": DATASET_DOI,
            "article_doi": ARTICLE_DOI,
            "download_url": DATASET_URL,
            "license": "CC-BY-4.0",
            "archive_filename": archive_path.name,
            "archive_size_bytes": archive_path.stat().st_size,
            "archive_md5": digest(archive_path, "md5"),
            "archive_sha256": digest(archive_path, "sha256"),
            "expected_md5": EXPECTED_MD5,
            "md5_matches_record": digest(archive_path, "md5") == EXPECTED_MD5,
        },
        "archive": {
            "member_count": len(archive_members),
            "uncompressed_size_bytes": sum(item["uncompressed_size"] for item in archive_members),
            "members": archive_members,
        },
        "extracted": {
            "file_count": len(file_records),
            "extension_counts": dict(sorted(extension_counts.items())),
            "candidate_numeric_2d_or_higher_structures": candidate_arrays,
            "files": file_records,
        },
        "automatic_decision": {
            "archive_integrity": "pass" if digest(archive_path, "md5") == EXPECTED_MD5 else "fail",
            "source_native_2d_candidate_count": len(candidate_arrays),
            "phase_2_authorized": False,
            "reason": "structural inventory requires scientific review before any field is admitted",
        },
    }


def render_markdown(summary: dict[str, Any]) -> str:
    dataset = summary["dataset"]
    extracted = summary["extracted"]
    candidates = extracted["candidate_numeric_2d_or_higher_structures"]
    lines = [
        "# R04 CdSeTe public archive structural audit",
        "",
        "## Provenance",
        "",
        f"- Dataset DOI: `{dataset['record_doi']}`",
        f"- Article DOI: `{dataset['article_doi']}`",
        f"- Download URL: `{dataset['download_url']}`",
        f"- License declared by record: `{dataset['license']}`",
        f"- Archive bytes: `{dataset['archive_size_bytes']}`",
        f"- MD5: `{dataset['archive_md5']}`",
        f"- SHA-256: `{dataset['archive_sha256']}`",
        f"- Zenodo MD5 match: `{dataset['md5_matches_record']}`",
        "",
        "## Structural inventory",
        "",
        f"- Archive members: `{summary['archive']['member_count']}`",
        f"- Extracted files: `{extracted['file_count']}`",
        f"- Uncompressed bytes: `{summary['archive']['uncompressed_size_bytes']}`",
        "- Extensions:",
    ]
    for suffix, count in extracted["extension_counts"].items():
        lines.append(f"  - `{suffix}`: `{count}`")
    lines.extend(["", "## Candidate numerical structures", ""])
    if candidates:
        for candidate in candidates:
            caution = f" — {candidate['caution']}" if "caution" in candidate else ""
            lines.append(
                f"- `{candidate['path']}`: shape `{candidate['shape']}`; "
                f"keyword hits `{candidate['keyword_hits']}`{caution}"
            )
    else:
        lines.append("No automatically detectable two-dimensional numerical structure was found.")
    lines.extend(
        [
            "",
            "## Decision boundary",
            "",
            "This report is structural only. It does not identify a physical observable, establish",
            "registration or coordinates, calibrate the native sample-plane kernel, or authorize",
            "Phase 2. A scientific review of the inventory and source article is required.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workdir", type=Path, default=Path("r04-cdsete-phase1"))
    parser.add_argument("--archive", type=Path, default=None)
    args = parser.parse_args()

    workdir = args.workdir.resolve()
    archive = args.archive.resolve() if args.archive else workdir / "Datasets.zip"
    extracted = workdir / "extracted"
    output = workdir / "audit"
    output.mkdir(parents=True, exist_ok=True)

    if not archive.exists():
        download(DATASET_URL, archive)

    archive_md5 = digest(archive, "md5")
    if archive_md5 != EXPECTED_MD5:
        raise RuntimeError(
            f"archive MD5 mismatch: expected {EXPECTED_MD5}, received {archive_md5}"
        )

    if extracted.exists():
        shutil.rmtree(extracted)
    archive_members = safe_extract(archive, extracted)
    file_records = [
        inspect_file(path, extracted)
        for path in sorted(extracted.rglob("*"))
        if path.is_file()
    ]
    summary = build_summary(archive, archive_members, file_records)

    inventory_path = output / "inventory.json"
    inventory_path.write_text(json.dumps(sanitize(summary), indent=2) + "\n", encoding="utf-8")
    report_path = output / "audit_summary.md"
    report_path.write_text(render_markdown(summary), encoding="utf-8")

    print(report_path.read_text(encoding="utf-8"))
    print(f"inventory: {inventory_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
